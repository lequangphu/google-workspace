# -*- coding: utf-8 -*-
"""Shared XLSX formatting utilities.

Consolidates duplicate formatting code from:
- generate_customers_xlsx.py
- generate_suppliers_xlsx.py
- generate_products_xlsx.py
"""

import logging
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

logger = logging.getLogger(__name__)

# Common styles used across all generators
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
DEFAULT_COLUMN_WIDTH = 20


class XLSXFormatter:
    """Shared XLSX formatting utilities for ERP export generators."""

    @staticmethod
    def format_header(
        worksheet,
        template: "ERPTemplate",
        column_width: int = DEFAULT_COLUMN_WIDTH,
    ) -> None:
        """Apply header styling and set column widths.

        Args:
            worksheet: openpyxl Worksheet to format
            template: ERPTemplate with COLUMN definitions
            column_width: Default column width
        """
        for col_idx, col_spec in enumerate(template.COLUMNS, start=1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, col_spec in enumerate(template.COLUMNS, start=1):
            letter = worksheet.cell(row=1, column=col_idx).column_letter
            worksheet.column_dimensions[letter].width = column_width

    @staticmethod
    def apply_column_formats(
        worksheet,
        template: "ERPTemplate",
        start_row: int = 2,
    ) -> None:
        """Apply number/date formats to data columns.

        Args:
            worksheet: openpyxl Worksheet to format
            template: ERPTemplate with COLUMN definitions
            start_row: First row of data (after header)
        """
        max_row = worksheet.max_row
        if max_row < start_row:
            return

        for col_idx, col_spec in enumerate(template.COLUMNS, start=1):
            if col_spec.format_code:
                for row in range(start_row, max_row + 1):
                    cell = worksheet.cell(row=row, column=col_idx)
                    cell.number_format = col_spec.format_code
                    if col_spec.data_type == "number":
                        cell.alignment = Alignment(horizontal="right")

    @staticmethod
    def write_xlsx(
        df: pd.DataFrame,
        output_path: Path,
        template: "ERPTemplate",
        sheet_name: str = "Sheet1",
        column_width: int = DEFAULT_COLUMN_WIDTH,
    ) -> None:
        """Write DataFrame to XLSX with standard formatting.

        Args:
            df: DataFrame to write
            output_path: Path to output XLSX file
            template: ERPTemplate with COLUMN definitions
            sheet_name: Name for the worksheet
            column_width: Default column width
        """
        output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name

        # Write headers
        for col_idx, col_spec in enumerate(template.COLUMNS, start=1):
            cell = worksheet.cell(row=1, column=col_idx, value=col_spec.name)

        # Write data
        for idx, row in df.iterrows():
            for col_idx, col_spec in enumerate(template.COLUMNS, start=1):
                value = row[col_spec.name]
                if pd.isna(value):
                    value = ""
                elif col_spec.data_type == "date":
                    if pd.notna(value) and isinstance(value, pd.Timestamp):
                        value = value.to_pydatetime()
                    elif pd.notna(value) and not isinstance(value, str):
                        value = str(value)
                elif not isinstance(value, str):
                    value = str(value)
                cell = worksheet.cell(row=idx + 2, column=col_idx, value=value)

        # Apply formatting
        XLSXFormatter.format_header(worksheet, template, column_width)
        XLSXFormatter.apply_column_formats(worksheet, template)

        workbook.save(output_path)
        logger.info(f"Wrote XLSX: {output_path}")

    @staticmethod
    def format_existing_xlsx(
        output_path: Path,
        template: "ERPTemplate",
        column_width: int = DEFAULT_COLUMN_WIDTH,
    ) -> None:
        """Apply formatting to an existing XLSX file.

        Args:
            output_path: Path to existing XLSX file
            template: ERPTemplate with COLUMN definitions
            column_width: Default column width
        """
        workbook = load_workbook(output_path)
        worksheet = workbook.active

        XLSXFormatter.format_header(worksheet, template, column_width)
        XLSXFormatter.apply_column_formats(worksheet, template)

        workbook.save(output_path)
        logger.info(f"Formatted XLSX: {output_path}")


def format_value(value, data_type: str) -> str:
    """Format value based on data type for Google Sheets upload.

    Args:
        value: Value to format
        data_type: One of 'text', 'number', 'date'

    Returns:
        Formatted value as string for Google Sheets
    """
    if value is None or value == "" or (isinstance(value, float) and pd.isna(value)):
        return ""

    if data_type == "number":
        if isinstance(value, (int, float)):
            return value
        try:
            cleaned = str(value).replace(",", "").replace(".", "").replace(" ", "")
            if cleaned.startswith("(") and cleaned.endswith(")"):
                cleaned = "-" + cleaned[1:-1]
            return float(cleaned)
        except (ValueError, TypeError):
            return str(value)

    elif data_type == "date":
        if isinstance(value, str):
            if "-" in value and len(value) >= 10:
                return value[:10]
            try:
                parsed = pd.to_datetime(value, errors="coerce")
                if pd.notna(parsed):
                    return parsed.strftime("%Y-%m-%d")
            except Exception:
                pass
        elif isinstance(value, pd.Timestamp):
            return value.strftime("%Y-%m-%d")
        return str(value)

    else:
        return str(value)
