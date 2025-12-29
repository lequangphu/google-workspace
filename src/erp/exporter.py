"""Export validated data to KiotViet XLSX files.

Maps validated DataFrames to KiotViet templates and writes XLSX files
with appropriate formatting and validation.
"""

import logging
from pathlib import Path
from typing import Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .templates import (
    CustomerTemplate,
    PriceBookTemplate,
    ProductTemplate,
)

logger = logging.getLogger(__name__)


def _load_inventory_with_latest_month(inventory_path: Path) -> pd.DataFrame:
    """Load inventory data and extract latest month's ending quantities and unit costs.

    Reads clean_inventory.py output (which contains multiple months) and
    extracts the "Số lượng cuối kỳ" (ending quantity) and "Đơn giá cuối kỳ"
    (ending unit price) from the latest month for each product.

    If the primary inventory_path is a validated summary, attempts to load
    the clean_inventory staging file instead.

    Args:
        inventory_path: Path to inventory CSV (validated or staging clean_inventory output)

    Returns:
        DataFrame with product codes, latest month's ending quantities, and unit costs
    """
    # First, try the provided path
    if inventory_path.exists():
        df = pd.read_csv(inventory_path)
        logger.info(f"Loaded inventory CSV with {len(df)} rows")

        # Check if this is output from clean_inventory.py (has date column)
        if "Ngày" in df.columns:
            logger.info("Processing clean_inventory output (multi-month data)")

            # Convert Ngày to datetime
            df["Ngày"] = pd.to_datetime(df["Ngày"], errors="coerce")

            # Sort by product and date
            df = df.sort_values(["Mã hàng", "Ngày"], na_position="last")

            # Extract latest month for each product
            latest_inventory = df.loc[df.groupby("Mã hàng")["Ngày"].idxmax()]

            # Create new DataFrame with mapping from Mã hàng to Số lượng cuối kỳ and Đơn giá cuối kỳ
            result = pd.DataFrame()
            result["Mã hàng"] = latest_inventory["Mã hàng"]
            result["Số lượng cuối kỳ"] = latest_inventory["Số lượng cuối kỳ"]
            if "Đơn giá cuối kỳ" in latest_inventory.columns:
                result["Đơn giá cuối kỳ"] = latest_inventory["Đơn giá cuối kỳ"]

            logger.info(
                f"Extracted latest month inventory for {len(result)} unique products"
            )
            return result
        else:
            logger.info(
                "Inventory CSV is a summary (no 'Ngày' column), "
                "attempting to load clean_inventory staging file..."
            )

            # Try to load from staging directory
            staging_clean_inventory = (
                inventory_path.parent.parent / "01-staging" / "import_export"
            )
            if staging_clean_inventory.exists():
                # Find the latest xuat_nhap_ton file (exclude adjustments)
                xnt_files = [
                    f
                    for f in staging_clean_inventory.glob("xuat_nhap_ton_*.csv")
                    if "_adjustments" not in f.name
                ]
                if xnt_files:
                    # Get the most recent file
                    latest_xnt = max(xnt_files, key=lambda p: p.stat().st_mtime)
                    logger.info(
                        f"Loading clean_inventory staging file: {latest_xnt.name}"
                    )

                    staging_df = pd.read_csv(latest_xnt)
                    if "Ngày" in staging_df.columns:
                        # Convert Ngày to datetime
                        staging_df["Ngày"] = pd.to_datetime(
                            staging_df["Ngày"], errors="coerce"
                        )

                        # Sort by product and date
                        staging_df = staging_df.sort_values(
                            ["Mã hàng", "Ngày"], na_position="last"
                        )

                        # Extract latest month for each product
                        latest_inventory = staging_df.loc[
                            staging_df.groupby("Mã hàng")["Ngày"].idxmax()
                        ]

                        # Create new DataFrame with mapping from Mã hàng to Số lượng cuối kỳ and Đơn giá cuối kỳ
                        result = pd.DataFrame()
                        result["Mã hàng"] = latest_inventory["Mã hàng"]
                        result["Số lượng cuối kỳ"] = latest_inventory[
                            "Số lượng cuối kỳ"
                        ]
                        if "Đơn giá cuối kỳ" in latest_inventory.columns:
                            result["Đơn giá cuối kỳ"] = latest_inventory[
                                "Đơn giá cuối kỳ"
                            ]

                        logger.info(
                            f"Extracted latest month inventory for {len(result)} unique products"
                        )
                        return result

            logger.warning(
                "Could not load clean_inventory staging file, using validated summary as-is"
            )
            return df
    else:
        logger.error(f"Inventory file not found: {inventory_path}")
        return pd.DataFrame()


def export_products_xlsx(
    product_info_path: Path,
    inventory_path: Path,
    price_sale_path: Path,
    enrichment_path: Path,
    output_path: Path,
) -> Tuple[Path, dict]:
    """Export product data to KiotViet Products XLSX.

    Merges product info, inventory, and pricing data, enriches from Google
    Sheet, maps to 9-column template, and exports as XLSX.

    Args:
        product_info_path: Path to product_info.csv
            (Mã hàng mới, Mã hàng, Tên hàng)
        inventory_path: Path to inventory.csv
            (Mã hàng mới, Tồn số lượng, Giá vốn FIFO)
        price_sale_path: Path to price_sale.csv
            (Mã hàng mới, Giá xuất cuối, ...)
        enrichment_path: Path to product enrichment CSV
            (Mã hàng, Nhóm hàng(3 Cấp), Thương hiệu)
        output_path: Path to write output XLSX

    Returns:
        Tuple of (output_path, stats_dict)
    """
    logger.info("=" * 70)
    logger.info("EXPORT PRODUCTS XLSX")
    logger.info("=" * 70)

    # Load all data
    logger.info(f"Loading product_info from {product_info_path}")
    products = pd.read_csv(product_info_path)

    logger.info(f"Loading inventory from {inventory_path}")
    inventory_latest_qty = _load_inventory_with_latest_month(inventory_path)

    # Also load validated inventory summary for cost columns
    logger.info(f"Loading validated inventory summary from {inventory_path}")
    inventory_summary = pd.read_csv(inventory_path)

    logger.info(f"Loading price_sale from {price_sale_path}")
    prices = pd.read_csv(price_sale_path)

    logger.info(f"Loading enrichment from {enrichment_path}")
    enrichment = pd.read_csv(enrichment_path)

    # Merge latest quantity and unit cost from clean_inventory
    # If unit cost not available from clean_inventory, fall back to validated summary
    if not inventory_latest_qty.empty:
        inventory = inventory_latest_qty.copy()

        # If Đơn giá cuối kỳ (unit cost) is present, rename to Giá vốn
        if "Đơn giá cuối kỳ" in inventory.columns:
            inventory["Giá vốn"] = inventory["Đơn giá cuối kỳ"]
            logger.info(
                "Using latest month's Đơn giá cuối kỳ as Giá vốn from clean_inventory"
            )
        else:
            # Fall back to validated summary for cost if not in clean_inventory
            if not inventory_summary.empty:
                inventory = inventory.merge(
                    inventory_summary[["Mã hàng", "Giá vốn"]],
                    on="Mã hàng",
                    how="left",
                )
                logger.info(
                    "Using Giá vốn from validated summary (Đơn giá cuối kỳ not in clean_inventory)"
                )
    else:
        # Fallback: use validated summary
        inventory = inventory_summary
        logger.info("Using validated inventory summary (clean_inventory not available)")

    # Merge data sources
    logger.info("Merging data sources...")
    df = products.copy()

    # Merge inventory by Mã hàng (from clean_inventory output)
    if "Mã hàng" in inventory.columns and "Mã hàng" in df.columns:
        df = df.merge(
            inventory,
            on="Mã hàng",
            how="left",
            suffixes=("", "_inventory"),
        )
        logger.info("Merged inventory by 'Mã hàng'")
    elif "Mã hàng mới" in inventory.columns:
        df = df.merge(
            inventory,
            on="Mã hàng mới",
            how="left",
            suffixes=("", "_inventory"),
        )
        logger.info("Merged inventory by 'Mã hàng mới'")
    else:
        logger.warning("Missing 'Mã hàng' or 'Mã hàng mới' in inventory data for merge")

    # Merge prices
    if "Mã hàng mới" in prices.columns:
        df = df.merge(
            prices,
            on="Mã hàng mới",
            how="left",
            suffixes=("", "_prices"),
        )
    else:
        logger.warning("Missing 'Mã hàng mới' in price_sale data")

    # Merge enrichment (by original product code or mã hàng)
    merge_key = None
    if "Mã hàng" in df.columns and "Mã hàng" in enrichment.columns:
        merge_key = "Mã hàng"
        df = df.merge(
            enrichment,
            on=merge_key,
            how="left",
            suffixes=("", "_enrichment"),
        )
    else:
        logger.warning(
            "Cannot merge enrichment: missing 'Mã hàng' in one or both sources"
        )

    logger.info(f"Merged data has {len(df)} products")

    # Build template DataFrame (27 columns)
    logger.info("Mapping to 27-column template...")
    template_data = pd.DataFrame(index=df.index)

    # Column 1: Loại hàng (constant)
    template_data["Loại hàng"] = "Hàng hóa"

    # Column 2: Nhóm hàng(3 Cấp) (from enrichment)
    if "Nhóm hàng(3 Cấp)" in df.columns:
        template_data["Nhóm hàng(3 Cấp)"] = df["Nhóm hàng(3 Cấp)"].fillna("")
    else:
        logger.warning("Missing 'Nhóm hàng(3 Cấp)' in enrichment")
        template_data["Nhóm hàng(3 Cấp)"] = ""

    # Column 3: Mã hàng (new product code)
    if "Mã hàng mới" in df.columns:
        template_data["Mã hàng"] = df["Mã hàng mới"]
    else:
        raise ValueError("Missing required column: Mã hàng mới")

    # Column 4: Mã vạch (original product code)
    if "Mã hàng" in df.columns:
        template_data["Mã vạch"] = df["Mã hàng"]
    else:
        template_data["Mã vạch"] = ""

    # Column 5: Tên hàng
    if "Tên hàng" in df.columns:
        template_data["Tên hàng"] = df["Tên hàng"]
    else:
        raise ValueError("Missing required column: Tên hàng")

    # Column 6: Thương hiệu (from enrichment)
    if "Thương hiệu" in df.columns:
        template_data["Thương hiệu"] = df["Thương hiệu"].fillna("")
    else:
        logger.warning("Missing 'Thương hiệu' in enrichment")
        template_data["Thương hiệu"] = ""

    # Column 7: Giá bán (max selling price from price_sale)
    if "Giá xuất cuối" in df.columns:
        template_data["Giá bán"] = pd.to_numeric(
            df["Giá xuất cuối"], errors="coerce"
        ).fillna(0)
    else:
        raise ValueError("Missing required column: Giá xuất cuối")

    # Column 8: Giá vốn (FIFO unit cost)
    cost_col = None
    for col in ["Giá vốn FIFO", "Giá vốn"]:
        if col in df.columns:
            cost_col = col
            break
    if cost_col:
        template_data["Giá vốn"] = pd.to_numeric(df[cost_col], errors="coerce").fillna(
            0
        )
    else:
        raise ValueError("Missing required column: Giá vốn")

    # Column 9: Tồn kho (current stock quantity from latest month)
    qty_col = None
    for col in ["Số lượng cuối kỳ", "Tồn số lượng", "Tổng số lượng"]:
        if col in df.columns:
            qty_col = col
            break
    if qty_col:
        template_data["Tồn kho"] = (
            pd.to_numeric(df[qty_col], errors="coerce").fillna(0).astype(int)
        )
    else:
        raise ValueError("Missing required column: Stock quantity")

    # Column 10: Tồn nhỏ nhất (min stock) - leave empty
    template_data["Tồn nhỏ nhất"] = ""

    # Column 11: Tồn lớn nhất (max stock) - leave empty
    template_data["Tồn lớn nhất"] = ""

    # Column 12: ĐVT (unit of measurement) - leave empty
    template_data["ĐVT"] = ""

    # Column 13: Mã ĐVT Cơ bản (base UOM code) - leave empty
    template_data["Mã ĐVT Cơ bản"] = ""

    # Column 14: Quy đổi (conversion factor) - leave empty
    template_data["Quy đổi"] = ""

    # Column 15: Thuộc tính (product attributes) - leave empty
    template_data["Thuộc tính"] = ""

    # Column 16: Mã HH Liên quan (related product IDs) - leave empty
    template_data["Mã HH Liên quan"] = ""

    # Column 17: Hình ảnh (url1,url2...) - leave empty
    template_data["Hình ảnh (url1,url2...)"] = ""

    # Column 18: Sử dụng Imei - leave empty
    template_data["Sử dụng Imei"] = ""

    # Column 19: Trọng lượng (weight) - leave empty
    template_data["Trọng lượng"] = ""

    # Column 20: Đang kinh doanh (active status) - default to 1 (active)
    template_data["Đang kinh doanh"] = 1

    # Column 21: Được bán trực tiếp (direct sale) - leave empty
    template_data["Được bán trực tiếp"] = ""

    # Column 22: Mô tả (description) - leave empty
    template_data["Mô tả"] = ""

    # Column 23: Mẫu ghi chú (note template) - leave empty
    template_data["Mẫu ghi chú"] = ""

    # Column 24: Vị trí (storage location) - leave empty
    template_data["Vị trí"] = ""

    # Column 25: Hàng thành phần (component products) - leave empty
    template_data["Hàng thành phần"] = ""

    # Column 26: Bảo hành (warranty) - leave empty
    template_data["Bảo hành"] = ""

    # Column 27: Bảo trì định kỳ (maintenance) - leave empty
    template_data["Bảo trì định kỳ"] = ""

    logger.info(f"Template DataFrame has {len(template_data)} rows, 27 columns")

    # Validate template
    logger.info("Validating template...")
    template = ProductTemplate()
    is_valid, errors = template.validate_dataframe(template_data)
    if not is_valid:
        logger.error(f"Template validation failed: {errors}")
        raise ValueError(f"Template validation failed: {errors}")
    logger.info("Template validation passed")

    # Write to XLSX
    output_path.parent.mkdir(parents=True, exist_ok=True)
    logger.info(f"Writing XLSX to {output_path}")
    template_data.to_excel(
        output_path, index=False, sheet_name="Sản phẩm", engine="openpyxl"
    )

    # Apply formatting
    logger.info("Applying formatting...")
    _format_product_xlsx(output_path, template)

    logger.info(f"Export complete: {output_path}")
    logger.info("=" * 70)

    stats = {
        "products_exported": len(template_data),
        "output_file": str(output_path),
    }
    return output_path, stats


def export_customers_xlsx(
    customer_ids_path: Path,
    enrichment_path: Path,
    output_path: Path,
) -> Tuple[Path, dict]:
    """Export customer data to KiotViet Customers XLSX.

    Args:
        customer_ids_path: Path to extract_customer_ids.csv
            (Mã khách hàng, Tên khách hàng, ...)
        enrichment_path: Path to customer enrichment CSV
        output_path: Path to write output XLSX

    Returns:
        Tuple of (output_path, stats_dict)
    """
    logger.info("=" * 70)
    logger.info("EXPORT CUSTOMERS XLSX")
    logger.info("=" * 70)

    # Load customer data
    logger.info(f"Loading customer data from {customer_ids_path}")
    customers = pd.read_csv(customer_ids_path)

    # Build template DataFrame (5 required columns minimum)
    logger.info("Mapping to customer template...")
    template_data = pd.DataFrame()

    # Column 1: Loại khách (constant - "Cá nhân")
    template_data["Loại khách"] = "Cá nhân"

    # Column 2: Mã khách hàng
    if "Mã khách hàng" in customers.columns:
        template_data["Mã khách hàng"] = customers["Mã khách hàng"]
    else:
        raise ValueError("Missing required column: Mã khách hàng")

    # Column 3: Tên khách hàng
    if "Tên khách hàng" in customers.columns:
        template_data["Tên khách hàng"] = customers["Tên khách hàng"]
    else:
        raise ValueError("Missing required column: Tên khách hàng")

    # Optional columns
    if "Điện thoại" in customers.columns:
        template_data["Điện thoại"] = customers["Điện thoại"].fillna("")
    if "Email" in customers.columns:
        template_data["Email"] = customers["Email"].fillna("")
    if "Ghi chú" in customers.columns:
        template_data["Ghi chú"] = customers["Ghi chú"].fillna("")

    logger.info(f"Customer DataFrame has {len(template_data)} rows")

    # Validate template
    logger.info("Validating template...")
    template = CustomerTemplate()
    is_valid, errors = template.validate_dataframe(template_data)
    if not is_valid:
        logger.error(f"Template validation failed: {errors}")
        raise ValueError(f"Template validation failed: {errors}")
    logger.info("Template validation passed")

    # Write to XLSX
    output_path.parent.mkdir(parents=True, exist_ok=True)
    logger.info(f"Writing XLSX to {output_path}")
    template_data.to_excel(
        output_path, index=False, sheet_name="Khách hàng", engine="openpyxl"
    )

    logger.info(f"Export complete: {output_path}")
    logger.info("=" * 70)

    stats = {
        "customers_exported": len(template_data),
        "output_file": str(output_path),
    }
    return output_path, stats


def export_pricebook_xlsx(
    product_info_path: Path,
    price_sale_path: Path,
    output_path: Path,
) -> Tuple[Path, dict]:
    """Export price book data to KiotViet PriceBook XLSX.

    Args:
        product_info_path: Path to product_info.csv
        price_sale_path: Path to price_sale.csv
        output_path: Path to write output XLSX

    Returns:
        Tuple of (output_path, stats_dict)
    """
    logger.info("=" * 70)
    logger.info("EXPORT PRICEBOOK XLSX")
    logger.info("=" * 70)

    # Load data
    logger.info(f"Loading product_info from {product_info_path}")
    products = pd.read_csv(product_info_path)

    logger.info(f"Loading price_sale from {price_sale_path}")
    prices = pd.read_csv(price_sale_path)

    # Merge
    logger.info("Merging data sources...")
    df = products.merge(prices, on="Mã hàng mới", how="left")

    # Build template DataFrame (5 columns)
    logger.info("Mapping to pricebook template...")
    template_data = pd.DataFrame()

    # Column 1: Mã hàng
    if "Mã hàng mới" in df.columns:
        template_data["Mã hàng"] = df["Mã hàng mới"]
    else:
        raise ValueError("Missing required column: Mã hàng mới")

    # Column 2: Tên hàng
    if "Tên hàng" in df.columns:
        template_data["Tên hàng"] = df["Tên hàng"]
    else:
        raise ValueError("Missing required column: Tên hàng")

    # Columns 3-5: Price lists (optional)
    if "Giá xuất cuối" in df.columns:
        template_data["Tên bảng giá 1"] = pd.to_numeric(
            df["Giá xuất cuối"], errors="coerce"
        ).fillna(0)
    else:
        template_data["Tên bảng giá 1"] = 0

    logger.info(f"PriceBook DataFrame has {len(template_data)} rows")

    # Validate template
    logger.info("Validating template...")
    template = PriceBookTemplate()
    is_valid, errors = template.validate_dataframe(template_data)
    if not is_valid:
        logger.error(f"Template validation failed: {errors}")
        raise ValueError(f"Template validation failed: {errors}")
    logger.info("Template validation passed")

    # Write to XLSX
    output_path.parent.mkdir(parents=True, exist_ok=True)
    logger.info(f"Writing XLSX to {output_path}")
    template_data.to_excel(
        output_path, index=False, sheet_name="Bảng giá", engine="openpyxl"
    )

    logger.info(f"Export complete: {output_path}")
    logger.info("=" * 70)

    stats = {
        "products_in_pricebook": len(template_data),
        "output_file": str(output_path),
    }
    return output_path, stats


def _format_product_xlsx(output_path: Path, template: ProductTemplate) -> None:
    """Apply Excel formatting to Products XLSX.

    Formats columns with appropriate alignment, number formats, and styles.

    Args:
        output_path: Path to XLSX file to format
        template: ProductTemplate instance with column specifications
    """
    workbook = load_workbook(output_path)
    worksheet = workbook.active

    # Header styling
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)

    # Apply header formatting
    for col_idx, col_spec in enumerate(template.COLUMNS, start=1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Apply column-specific formatting
    for col_idx, col_spec in enumerate(template.COLUMNS, start=1):
        # Set column width
        worksheet.column_dimensions[
            worksheet.cell(row=1, column=col_idx).column_letter
        ].width = 18

        # Apply data formatting
        if col_spec.format_code:
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=col_idx)
                cell.number_format = col_spec.format_code
                cell.alignment = Alignment(horizontal="right")
        else:
            # Text alignment (left)
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=col_idx)
                cell.alignment = Alignment(horizontal="left")

    workbook.save(output_path)
