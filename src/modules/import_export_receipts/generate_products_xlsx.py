# -*- coding: utf-8 -*-
"""Generate Products data directly from staging data.

This module consolidates extract_products.py and exporter.py to:
1. Read product codes from chi_tiet_nhap_cleaned.csv
2. Get Giá bán (max selling price) from chi_tiet_xuat_cleaned.csv
3. Get Giá vốn and Tồn kho from clean_inventory.py output (Đơn giá cuối kỳ, Số lượng cuối kỳ)
4. Enrich with Nhóm hàng, Thương hiệu, Tên hàng from Google Sheets
5. Write to Google Spreadsheet (Sản phẩm tab) and optionally to data/03-erp-export/Products.xlsx

Raw source: import_export_receipts
Module: import_export_receipts
Pipeline stage: data/01-staging/ → Google Spreadsheet
"""

import logging
import toml
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.erp.templates import ProductTemplate
from src.utils.product_attributes import extract_attributes
from src.utils.staging_cache import StagingCache

logger = logging.getLogger(__name__)

WORKSPACE_ROOT = Path(__file__).parent.parent.parent
PIPELINE_CONFIG = WORKSPACE_ROOT / "pipeline.toml"

_CONFIG = toml.load(PIPELINE_CONFIG) if PIPELINE_CONFIG.exists() else {}

CONFIG = {
    "staging_dir": Path.cwd() / "data" / "01-staging" / "import_export",
    "export_dir": Path.cwd() / "data" / "03-erp-export",
    "nhap_pattern": "*Chi tiết nhập*.csv",
    "xuat_pattern": "*Chi tiết xuất*.csv",
    "xnt_pattern": "Xuất nhập tồn *.csv",
    "product_lookup_spreadsheet_id": _CONFIG.get("enrichment", {}).get(
        "product_lookup_spreadsheet_id", "16bGN2gjWspCqlFD4xB--7WtkYtTpDaWzRQx9sV97ed8"
    ),
    "product_lookup_sheet_name": "Nhóm hàng, thương hiệu",
    "reports_spreadsheet_id": _CONFIG.get("upload", {}).get(
        "google_sheets_id_reports", "11vk-p0iL9JcNH180n4uV5VTuPnhJ97lBgsLEfCnWx_k"
    ),
    "products_sheet_name": "products_to_import",
}


def find_latest_file(directory: Path, pattern: str) -> Optional[Path]:
    """Find the latest file matching pattern in directory."""
    matching_files = list(directory.glob(pattern))
    if not matching_files:
        raise FileNotFoundError(f"No files matching {pattern} found in {directory}")
    latest_file = sorted(matching_files, key=lambda p: p.stat().st_mtime)[-1]
    logger.info(f"Found input file: {latest_file.name}")
    return latest_file


def get_latest_inventory(xnt_dir: Path) -> pd.DataFrame:
    """Load clean_inventory output and extract latest month for each product.

    Returns DataFrame with columns: Mã hàng, Số lượng cuối kỳ, Đơn giá cuối kỳ
    """
    xnt_files = list(xnt_dir.glob(CONFIG["xnt_pattern"]))
    xnt_files = [f for f in xnt_files if "adjustments" not in f.name.lower()]

    if not xnt_files:
        raise FileNotFoundError(f"No xuat_nhap_ton_*.csv files found in {xnt_dir}")

    all_data = []
    for f in xnt_files:
        df = StagingCache.get_dataframe(f)
        if "Ngày" in df.columns:
            all_data.append(df)

    if not all_data:
        raise ValueError("No XNT files contain Ngày column")

    combined = pd.concat(all_data, ignore_index=True)
    combined["Ngày"] = pd.to_datetime(combined["Ngày"], errors="coerce")

    latest = combined.loc[combined.groupby("Mã hàng")["Ngày"].idxmax()]
    result = pd.DataFrame(
        {
            "Mã hàng": latest["Mã hàng"],
            "Số lượng cuối kỳ": pd.to_numeric(
                latest["Số lượng cuối kỳ"], errors="coerce"
            ).clip(lower=0),
            "Đơn giá cuối kỳ": pd.to_numeric(
                latest["Đơn giá cuối kỳ"], errors="coerce"
            ),
        }
    )

    logger.info(f"Extracted latest inventory for {len(result)} products")
    return result


def fetch_product_lookup() -> pd.DataFrame:
    """Fetch product lookup from Google Sheets."""
    try:
        from src.modules.google_api import connect_to_drive, read_sheet_data
    except ImportError:
        logger.error("Cannot import google_api. Skipping product lookup.")
        return pd.DataFrame()

    try:
        logger.info("Fetching product lookup from Google Sheets...")
        _, sheets_service = connect_to_drive()

        values = read_sheet_data(
            sheets_service,
            CONFIG["product_lookup_spreadsheet_id"],
            CONFIG["product_lookup_sheet_name"],
        )

        if not values or len(values) < 2:
            logger.warning("Product lookup sheet is empty")
            return pd.DataFrame()

        headers = values[0]
        data_rows = [
            row[: len(headers)] + [""] * max(0, len(headers) - len(row))
            for row in values[1:]
        ]

        df = pd.DataFrame(data_rows, columns=headers)
        required = ["Mã hàng", "Nhóm hàng(3 Cấp)", "Thương hiệu"]
        missing = [c for c in required if c not in df.columns]
        if missing:
            logger.error(f"Product lookup missing columns: {missing}")
            return pd.DataFrame()

        for col in required + (["Tên hàng"] if "Tên hàng" in df.columns else []):
            df[col] = df[col].astype(str).str.strip().replace(["", "None"], pd.NA)

        df = df.drop_duplicates(subset=["Mã hàng"], keep="first")
        logger.info(f"Fetched {len(df)} products from lookup")
        return df

    except Exception as e:
        logger.error(f"Failed to fetch product lookup: {e}")
        return pd.DataFrame()


def standardize_brand_names(enrichment_df: pd.DataFrame) -> pd.DataFrame:
    """Standardize brand names in Tên hàng column."""
    replacements = {
        "chengsin": "CHENGSHIN",
        "michenlin": "MICHELIN",
        "caosumina": "CASUMINA",
    }

    if "Tên hàng" not in enrichment_df.columns:
        return enrichment_df

    df = enrichment_df.copy()
    for old, new in replacements.items():
        df["Tên hàng"] = df["Tên hàng"].str.replace(old, new, case=False, regex=False)
    return df


def get_product_names_from_nhap(nhap_df: pd.DataFrame) -> pd.DataFrame:
    """Extract unique product names from nhap data, choosing shortest name per product.

    This is used as fallback when Google Sheets Tên hàng is null/empty.

    Args:
        nhap_df: DataFrame with 'Mã hàng' and 'Tên hàng' columns

    Returns:
        DataFrame with 'Mã hàng' and 'Tên hàng' columns (shortest name per product)
    """
    if nhap_df.empty or "Tên hàng" not in nhap_df.columns:
        return pd.DataFrame(columns=["Mã hàng", "Tên hàng"])

    nhap_df = nhap_df.copy()

    nhap_df["Tên hàng"] = (
        nhap_df["Tên hàng"]
        .astype(str)
        .str.strip()
        .replace(["nan", "None", "<NA>", ""], pd.NA)
    )

    nhap_df = nhap_df.dropna(subset=["Tên hàng"])

    if nhap_df.empty:
        return pd.DataFrame(columns=["Mã hàng", "Tên hàng"])

    def get_shortest_name(group):
        names = group["Tên hàng"].dropna().astype(str)
        if names.empty:
            return ""
        return min(names, key=len)

    product_names = nhap_df.groupby("Mã hàng", as_index=False, group_keys=False).apply(
        get_shortest_name, include_groups=False
    )
    product_names.columns = ["Mã hàng", "Tên hàng"]

    logger.info(f"Extracted {len(product_names)} product names from nhap data")
    return product_names


def get_sort_key(ma_hang: str, nhap_df: pd.DataFrame) -> tuple:
    """Get sort key based on first occurrence date in nhập data."""
    product_data = nhap_df[nhap_df["Mã hàng"] == ma_hang]
    if product_data.empty:
        return (pd.NaT, "", 0)

    first_row = product_data.iloc[0]
    try:
        date_val = pd.to_datetime(first_row.get("Ngày", ""), errors="coerce")
    except Exception:
        date_val = pd.NaT

    return (
        date_val,
        str(first_row.get("Mã chứng từ", "")),
        float(first_row.get("Thành tiền", 0) or 0),
    )


def calculate_max_selling_price(xuat_df: pd.DataFrame) -> pd.DataFrame:
    """Calculate max selling price per product."""
    if xuat_df.empty:
        return pd.DataFrame()

    df = xuat_df.copy()
    df["Đơn giá"] = df["Thành tiền"] / df["Số lượng"]
    df["Đơn giá"] = df["Đơn giá"].replace([float("inf"), float("-inf")], 0)

    max_prices = df.groupby("Mã hàng")["Đơn giá"].max().reset_index()
    max_prices.columns = ["Mã hàng", "Giá bán"]

    logger.info(f"Calculated max selling prices for {len(max_prices)} products")
    return max_prices


def build_template_dataframe(
    product_codes: pd.DataFrame,
    inventory: pd.DataFrame,
    prices: pd.DataFrame,
    enrichment: pd.DataFrame,
    fallback_names: Optional[pd.DataFrame] = None,
) -> pd.DataFrame:
    """Build 27-column KiotViet template DataFrame.

    Args:
        product_codes: DataFrame with Mã hàng and Mã hàng mới
        inventory: DataFrame with Số lượng cuối kỳ, Đơn giá cuối kỳ
        prices: DataFrame with Giá bán
        enrichment: DataFrame with Nhóm hàng(3 Cấp), Thương hiệu, Tên hàng, Loại máy xe, Tên xe, Thuộc tính, Mô tả from Google Sheets
        fallback_names: Optional DataFrame with Mã hàng, Tên hàng from nhap data (shortest name)
    """
    df = product_codes.copy()

    df = df.merge(inventory, on="Mã hàng", how="left")

    df = df.merge(prices, on="Mã hàng", how="left")

    if not enrichment.empty:
        enrich_cols = ["Mã hàng", "Nhóm hàng(3 Cấp)", "Thương hiệu"]
        if "Tên hàng" in enrichment.columns:
            enrich_cols.append("Tên hàng")
        if "Loại máy xe" in enrichment.columns:
            enrich_cols.append("Loại máy xe")
        if "Tên xe" in enrichment.columns:
            enrich_cols.append("Tên xe")
        if "Thuộc tính" in enrichment.columns:
            enrich_cols.append("Thuộc tính")
        if "Mô tả" in enrichment.columns:
            enrich_cols.append("Mô tả")
        df = df.merge(enrichment[enrich_cols], on="Mã hàng", how="left")

    # Fill fallback values for enrichment columns
    df["Nhóm hàng(3 Cấp)"] = df.get("Nhóm hàng(3 Cấp)", pd.Series(dtype=str)).fillna(
        "Phụ tùng khác>>Xe khác"
    )
    df["Thương hiệu"] = df.get("Thương hiệu", pd.Series(dtype=str)).fillna(
        "Chưa có thương hiệu"
    )
    df["Loại máy xe"] = df.get("Loại máy xe", pd.Series(dtype=str)).fillna("")
    df["Tên xe"] = df.get("Tên xe", pd.Series(dtype=str)).fillna("")

    if fallback_names is not None and not fallback_names.empty:
        if "Tên hàng" in df.columns:
            fallback_map = fallback_names.set_index("Mã hàng")["Tên hàng"].to_dict()

            df["Tên hàng"] = df["Tên hàng"].fillna(pd.NA).replace("", pd.NA)

            df["Tên hàng"] = df.apply(
                lambda row: (
                    row["Tên hàng"]
                    if pd.notna(row["Tên hàng"]) and row["Tên hàng"] != ""
                    else fallback_map.get(row["Mã hàng"], "")
                ),
                axis=1,
            )

            fallback_count = (df["Tên hàng"].isin(fallback_map.values())).sum()
            logger.info(f"Applied fallback product names for {fallback_count} products")

    template = ProductTemplate()
    result = pd.DataFrame(index=df.index)

    col_mapping = {
        "Loại hàng": ("Hàng hóa", None),
        "Nhóm hàng(3 Cấp)": (df.get("Nhóm hàng(3 Cấp)", ""), ""),
        "Mã hàng": (df["Mã hàng mới"], None),
        "Mã vạch": ("", ""),
        "Tên hàng": (df.get("Tên hàng", ""), ""),
        "Thương hiệu": (df.get("Thương hiệu", ""), ""),
        "Giá bán": (
            pd.to_numeric(df.get("Giá bán", 0), errors="coerce").fillna(0),
            "#,0.##0",
        ),
        "Giá vốn": (
            pd.to_numeric(df.get("Đơn giá cuối kỳ", 0), errors="coerce").fillna(0),
            "#,0.##0",
        ),
        "Tồn kho": (
            pd.to_numeric(df.get("Số lượng cuối kỳ", 0), errors="coerce")
            .fillna(0)
            .astype(int),
            "#,0.##0",
        ),
    }

    for col_name, (data, format_code) in col_mapping.items():
        result[col_name] = data

    empty_cols = [
        "Tồn nhỏ nhất",
        "Tồn lớn nhất",
        "ĐVT",
        "Mã ĐVT Cơ bản",
        "Quy đổi",
        "Mã HH Liên quan",
        "Hình ảnh (url1,url2...)",
        "Sử dụng Imei",
        "Trọng lượng",
        "Được bán trực tiếp",
        "Mô tả",
        "Mẫu ghi chú",
        "Vị trí",
        "Hàng thành phần",
        "Bảo hành",
        "Bảo trì định kỳ",
    ]
    for col in empty_cols:
        if col == "Thuộc tính" and "Thuộc tính" in df.columns:
            result[col] = df["Thuộc tính"]
        elif col == "Mô tả" and "Mô tả" in df.columns:
            result[col] = df["Mô tả"]
        else:
            result[col] = ""

    result["Đang kinh doanh"] = 1

    is_valid, errors = template.validate_dataframe(result)
    if not is_valid:
        logger.error(f"Template validation failed: {errors}")
        raise ValueError(f"Template validation failed: {errors}")

    logger.info(f"Template DataFrame: {len(result)} rows, 27 columns")
    return result


def format_product_xlsx(output_path: Path) -> None:
    """Apply Excel formatting to Products XLSX."""
    workbook = load_workbook(output_path)
    worksheet = workbook.active

    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)

    template = ProductTemplate()

    for col_idx, col_spec in enumerate(template.COLUMNS, start=1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, col_spec in enumerate(template.COLUMNS, start=1):
        letter = worksheet.cell(row=1, column=col_idx).column_letter
        worksheet.column_dimensions[letter].width = 18

        if col_spec.format_code:
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=col_idx)
                cell.number_format = col_spec.format_code
                cell.alignment = Alignment(horizontal="right")
        else:
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=col_idx)
                cell.alignment = Alignment(horizontal="left")

    workbook.save(output_path)


def write_to_spreadsheet(
    df: pd.DataFrame,
    spreadsheet_id: str,
    sheet_name: str,
    raw_columns: list[int] = None,
) -> bool:
    """Write DataFrame to Google Spreadsheet as a sheet tab.

    Args:
        df: DataFrame to write
        spreadsheet_id: Google Spreadsheet ID
        sheet_name: Name of the sheet tab
        raw_columns: Optional list of column indices to upload as RAW

    Returns:
        True if successful, False otherwise
    """
    try:
        from src.modules.google_api import connect_to_drive, upload_dataframe_to_sheet
    except ImportError:
        logger.error("Cannot import google_api. Cannot write to spreadsheet.")
        return False

    try:
        logger.info(f"Writing to Google Spreadsheet: {sheet_name}")
        _, sheets_service = connect_to_drive()

        values = [df.columns.tolist()] + df.values.tolist()
        values = [["" if pd.isna(v) else v for v in row] for row in values]

        success = upload_dataframe_to_sheet(
            sheets_service,
            spreadsheet_id,
            sheet_name,
            df,
            raw_columns=raw_columns,
        )
        if success:
            logger.info(f"Successfully wrote {len(df)} rows to '{sheet_name}' sheet")
        return success

    except Exception as e:
        logger.error(f"Failed to write to spreadsheet: {e}")
        return False


def process(
    staging_dir: Optional[Path] = None, write_to_sheets: bool = True
) -> Optional[Path]:
    """Generate Products.xlsx from staging data.

    Args:
        staging_dir: Directory with staged receipt data (defaults to config)

    Returns:
        Path to Products.xlsx or None if failed
    """
    if staging_dir is None:
        staging_dir = CONFIG["staging_dir"]

    if not staging_dir.exists():
        logger.error(f"Staging directory not found: {staging_dir}")
        return None

    logger.info("=" * 70)
    logger.info("GENERATING PRODUCTS.XLSX")
    logger.info("=" * 70)

    try:
        nhap_file = find_latest_file(staging_dir, CONFIG["nhap_pattern"])
        nhap_df = StagingCache.get_dataframe(nhap_file)
        logger.info(f"Loaded nhập data: {len(nhap_df)} rows")

        xuat_file = find_latest_file(staging_dir, CONFIG["xuat_pattern"])
        xuat_df = StagingCache.get_dataframe(xuat_file)
        logger.info(f"Loaded xuất data: {len(xuat_df)} rows")

        inventory_df = get_latest_inventory(staging_dir)
        logger.info(f"Loaded inventory data: {len(inventory_df)} products")

        CUTOFF_DATE = pd.Timestamp("2025-01-01")

        xuat_df["Ngày"] = pd.to_datetime(xuat_df["Ngày"], errors="coerce")
        products_with_sales = set(
            xuat_df[xuat_df["Ngày"] >= CUTOFF_DATE]["Mã hàng"].unique()
        )

        products_with_inventory = set(
            inventory_df[inventory_df["Số lượng cuối kỳ"] > 0]["Mã hàng"].unique()
        )

        valid_products = products_with_sales | products_with_inventory
        logger.info(
            f"Filtered to {len(valid_products)} products (sales since 2025-01-01 or inventory > 0)"
        )

        prices_df = calculate_max_selling_price(xuat_df)

        # Note: fetch_product_lookup() calls Google Sheets API directly, bypassing StagingCache.
        # This is intentional - lookup data is external enrichment data fetched on-demand,
        # not a local CSV file in the pipeline stages that benefits from mtime-based caching.
        lookup_df = fetch_product_lookup()

        if lookup_df.empty:
            logger.warning("Product lookup empty, using placeholders")
            enrichment_df = pd.DataFrame(
                {
                    "Mã hàng": nhap_df["Mã hàng"].unique(),
                    "Nhóm hàng(3 Cấp)": "Chưa có nhóm hàng",
                    "Thương hiệu": "Chưa có thương hiệu",
                    "Tên hàng": "",
                }
            )
        else:
            enrichment_df = lookup_df.copy()

        enrichment_df = standardize_brand_names(enrichment_df)

        if "Tên hàng" in enrichment_df.columns:
            enrichment_df["Thuộc tính"] = enrichment_df["Tên hàng"].apply(
                extract_attributes
            )
        else:
            enrichment_df["Thuộc tính"] = ""

        matched = enrichment_df["Thuộc tính"].ne("").sum()
        total = len(enrichment_df)
        logger.info(f"Extracted attributes for {matched}/{total} products")

        unique_products = nhap_df[nhap_df["Mã hàng"].isin(valid_products)][
            "Mã hàng"
        ].unique()

        sort_keys = {
            ma_hang: get_sort_key(ma_hang, nhap_df) for ma_hang in unique_products
        }
        sorted_products = sorted(
            unique_products, key=lambda x: sort_keys.get(x, (pd.NaT, "", 0))
        )

        product_codes = pd.DataFrame(
            {
                "Mã hàng": sorted_products,
                "Mã hàng mới": ["SPC" + str(m) for m in sorted_products],
            }
        )
        logger.info(f"Generated product codes for {len(product_codes)} products")

        fallback_names = get_product_names_from_nhap(nhap_df)

        template_df = build_template_dataframe(
            product_codes, inventory_df, prices_df, enrichment_df, fallback_names
        )

        if write_to_sheets:
            success = write_to_spreadsheet(
                template_df,
                CONFIG["reports_spreadsheet_id"],
                CONFIG["products_sheet_name"],
                raw_columns=[2],
            )
            if not success:
                logger.warning("Failed to write to spreadsheet, continuing anyway...")

        CONFIG["export_dir"].mkdir(parents=True, exist_ok=True)
        output_path = CONFIG["export_dir"] / "Products.xlsx"

        template_df.to_excel(
            output_path, index=False, sheet_name="Sản phẩm", engine="openpyxl"
        )
        logger.info(f"Wrote template to {output_path}")

        format_product_xlsx(output_path)
        logger.info("Formatted Products.xlsx")

        logger.info("=" * 70)
        logger.info(f"PRODUCTS.XLSX GENERATED: {output_path}")
        logger.info(f"Total products: {len(template_df)}")
        logger.info("=" * 70)

        return output_path

    except Exception as e:
        logger.error(f"Products.xlsx generation failed: {e}", exc_info=True)
        raise


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
    )
    process()
