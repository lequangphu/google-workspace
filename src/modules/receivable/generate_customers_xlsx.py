# -*- coding: utf-8 -*-
"""Generate KiotViet customer import XLSX and upload to Google Sheets.

This single script:
1. Reads customer data from Google Sheets (Thong tin KH, TỔNG CÔNG NỢ)
2. Reads transaction data from staged clean_receipts_sale_*.csv
3. Merges and transforms data in-memory
4. Generates unified customer codes (KH000001, KH000002, ...)
5. Maps to KiotViet 20-column template
6. Exports to data/03-erp-export/Customers.xlsx
7. Uploads to Google Sheet as "customers_to_import" tab

Raw sources:
- Google Sheets: 1kouZwJy8P_zZhjjn49Lfbp3KN81mhHADV7VKDhv5xkM (Thong tin KH, TỔNG CÔNG NỢ)
- Local CSVs: data/01-staging/import_export/clean_receipts_sale_*.csv

Output:
- XLSX: data/03-erp-export/Customers.xlsx
- Google Sheet: 11vk-p0iL9JcNH180n4uV5VTuPnhJ97lBgsLEfCnWx_k (tab: customers_to_import)
"""

import logging
from pathlib import Path
from typing import Optional

import pandas as pd
import toml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.erp.templates import CustomerTemplate
from src.modules.google_api import (
    authenticate_google,
    read_sheet_data,
    upload_dataframe_to_sheet,
)
from src.utils.data_cleaning import (
    generate_entity_codes,
    merge_master_data,
    parse_numeric,
    split_phone_numbers,
)
from src.utils.staging_cache import StagingCache

logger = logging.getLogger(__name__)


DATA_STAGING_DIR = Path.cwd() / "data" / "01-staging"
IMPORT_EXPORT_STAGING = DATA_STAGING_DIR / "import_export"
DATA_EXPORT_DIR = Path.cwd() / "data" / "03-erp-export"
PIPELINE_CONFIG = Path.cwd() / "pipeline.toml"

_CONFIG = toml.load(PIPELINE_CONFIG) if PIPELINE_CONFIG.exists() else {}
RECEIVABLE_SPREADSHEET_ID = (
    _CONFIG.get("sources", {})
    .get("receivable", {})
    .get("spreadsheet_id", "1kouZwJy8P_zZhjjn49Lfbp3KN81mhHADV7VKDhv5xkM")
)
EXPORT_SPREADSHEET_ID = _CONFIG.get("upload", {}).get(
    "google_sheets_id_reports", "11vk-p0iL9JcNH180n4uV5VTuPnhJ97lBgsLEfCnWx_k"
)
EXPORT_SHEET_NAME = "customers_to_import"


def load_config() -> dict:
    """Load configuration from pipeline.toml."""
    if not PIPELINE_CONFIG.exists():
        logger.warning(f"Config file not found: {PIPELINE_CONFIG}")
        return {}

    try:
        config = toml.load(PIPELINE_CONFIG)
        return config
    except Exception as e:
        logger.error(f"Failed to load config: {e}")
        return {}


def read_google_sheet(sheets_service, spreadsheet_id: str, sheet_name: str) -> list:
    """Read data from a Google Sheet tab."""
    raw_data = read_sheet_data(sheets_service, spreadsheet_id, sheet_name)
    if not raw_data:
        logger.warning(f"No data from {sheet_name}")
        return []

    header_row_idx = None
    for idx, row in enumerate(raw_data):
        if row and str(row[0]).strip() == "STT":
            header_row_idx = idx
            break

    if header_row_idx is None:
        logger.warning(f"Could not find header row in {sheet_name}")
        return []

    header_row = raw_data[header_row_idx]
    data_rows = raw_data[header_row_idx + 1 :]

    max_cols = len(header_row)
    padded_rows = []
    for row in data_rows:
        if len(row) < max_cols:
            row = row + [""] * (max_cols - len(row))
        padded_rows.append(row[:max_cols])

    return [header_row] + padded_rows


def load_thong_tin_kh(sheets_service) -> pd.DataFrame:
    """Load and clean customer info from Thong tin KH tab."""
    logger.info("Loading Thong tin KH...")
    raw_data = read_google_sheet(
        sheets_service, RECEIVABLE_SPREADSHEET_ID, "Thong tin KH"
    )
    if not raw_data:
        return pd.DataFrame()

    header = raw_data[0]
    rows = raw_data[1:]

    col_map = {}
    for idx, col in enumerate(header):
        col_clean = col.strip() if col else ""
        if col_clean == "MÃ KH":
            col_map["MÃ KH"] = idx
        elif col_clean == "TÊN KHÁCH HÀNG":
            col_map["TÊN KHÁCH HÀNG"] = idx
        elif col_clean == "Địa chỉ":
            col_map["Địa chỉ"] = idx
        elif col_clean == "Tel":
            col_map["Tel"] = idx
        elif col_clean == "Ghi chú":
            col_map["Ghi chú"] = idx

    if "TÊN KHÁCH HÀNG" not in col_map:
        logger.warning("Required column TÊN KHÁCH HÀNG not found")
        return pd.DataFrame()

    data = []
    for row in rows:
        if len(row) <= col_map.get("TÊN KHÁCH HÀNG", 0):
            continue

        name = str(row[col_map["TÊN KHÁCH HÀNG"]]).strip()
        if not name:
            continue

        if name.isdigit():
            continue

        if name.upper() in ("TỔNG CỘNG", "NL", "NGƯỜI LẬP", "TYPO"):
            continue

        ma_kh_col = col_map.get("MÃ KH")
        ma_kh_val = ""
        if ma_kh_col is not None and len(row) > ma_kh_col:
            ma_kh_val = str(row[ma_kh_col]).strip()

        if not ma_kh_val:
            continue

        tel_col = col_map.get("Tel")
        tel_raw = ""
        if tel_col is not None and len(row) > tel_col:
            tel_raw = str(row[tel_col]).strip()

        tel_list = split_phone_numbers(tel_raw)
        phone_value = tel_list[0] if tel_list else ""

        data.append(
            {
                "Mã KH cũ": ma_kh_val,
                "Tên khách hàng": name,
                "Điện thoại": phone_value,
                "Địa chỉ": str(row[col_map.get("Địa chỉ", -1)]).strip()
                if col_map.get("Địa chỉ") is not None and len(row) > col_map["Địa chỉ"]
                else "",
                "Ghi chú": str(row[col_map.get("Ghi chú", -1)]).strip()
                if col_map.get("Ghi chú") is not None and len(row) > col_map["Ghi chú"]
                else "",
            }
        )

    df = pd.DataFrame(data)
    logger.info(f"Loaded {len(df)} customers from Thong tin KH")
    return df


def load_tong_cong_no(sheets_service) -> pd.DataFrame:
    """Load and clean debt data from TỔNG CÔNG NỢ tab."""
    logger.info("Loading TỔNG CÔNG NỢ...")
    raw_data = read_google_sheet(
        sheets_service, RECEIVABLE_SPREADSHEET_ID, "TỔNG CÔNG NỢ"
    )
    if not raw_data:
        return pd.DataFrame()

    header = raw_data[0]
    rows = raw_data[1:]

    col_map = {}
    for idx, col in enumerate(header):
        col_clean = col.strip() if col else ""
        if col_clean == "TÊN KHÁCH HÀNG":
            col_map["TÊN KHÁCH HÀNG"] = idx
        elif "NỢ CÒN LẠI" in col_clean:
            col_map["NỢ CÒN LẠI"] = idx

    if "TÊN KHÁCH HÀNG" not in col_map:
        logger.warning("Required column TÊN KHÁCH HÀNG not found in TỔNG CÔNG NỢ")
        return pd.DataFrame()

    data = []
    for row in rows:
        if len(row) <= col_map.get("TÊN KHÁCH HÀNG", 0):
            continue

        name = str(row[col_map["TÊN KHÁCH HÀNG"]]).strip()
        if not name:
            continue

        if name.isdigit():
            continue

        if name.upper() in ("TỔNG CỘNG", "NL", "NGƯỜI LẬP", "TYPO", "TRUNGKL"):
            continue

        debt_col = col_map.get("NỢ CÒN LẠI")
        debt_val = "0"
        if debt_col is not None and len(row) > debt_col:
            debt_val = parse_numeric(row[debt_col])

        data.append(
            {
                "Tên khách hàng": name,
                "Nợ cần thu hiện tại": debt_val,
            }
        )

    df = pd.DataFrame(data)
    df["Nợ cần thu hiện tại"] = pd.to_numeric(
        df["Nợ cần thu hiện tại"], errors="coerce"
    ).fillna(0)
    logger.info(f"Loaded {len(df)} debt records from TỔNG CÔNG NỢ")
    return df


def load_sale_transactions() -> pd.DataFrame:
    """Load cleaned sale transactions from staging directory."""
    logger.info("Loading sale transactions from staging...")

    receipt_files = list(IMPORT_EXPORT_STAGING.glob("Chi tiết xuất*.csv"))
    if not receipt_files:
        logger.warning("No Chi tiết xuất*.csv files found")
        return pd.DataFrame()

    all_data = []
    for f in receipt_files:
        try:
            df = StagingCache.get_dataframe(f)
            all_data.append(df)
            logger.info(f"  Loaded {len(df)} rows from {f.name}")
        except Exception as e:
            logger.warning(f"Error reading {f}: {e}")

    if not all_data:
        return pd.DataFrame()

    combined = pd.concat(all_data, ignore_index=True)
    logger.info(f"Total transactions: {len(combined)}")
    return combined


def aggregate_transactions(df: pd.DataFrame) -> pd.DataFrame:
    """Aggregate transactions by customer name."""
    if df.empty or "Tên khách hàng" not in df.columns:
        return pd.DataFrame()

    logger.info("Aggregating transactions by customer...")

    df = df.copy()

    if "Ngày" in df.columns:
        df["Ngày"] = pd.to_datetime(df["Ngày"], errors="coerce")

    amount_col = "Thành tiền" if "Thành tiền" in df.columns else "Số lượng"

    aggregated = (
        df.groupby("Tên khách hàng", dropna=False)
        .agg(
            first_date=("Ngày", "min"),
            last_date=("Ngày", "max"),
            total_amount=(amount_col, "sum"),
            transaction_count=("Ngày", "count"),
        )
        .reset_index()
    )

    aggregated = aggregated.dropna(subset=["Tên khách hàng"])
    aggregated["Tên khách hàng"] = aggregated["Tên khách hàng"].str.strip()

    logger.info(f"Aggregated {len(aggregated)} unique customers")
    return aggregated


def merge_all_data(
    customers_df: pd.DataFrame,
    debts_df: pd.DataFrame,
    transactions_df: pd.DataFrame,
) -> pd.DataFrame:
    """Merge all data sources by customer name."""
    return merge_master_data(
        master_df=customers_df,
        debts_df=debts_df,
        transactions_df=transactions_df,
        name_column="Tên khách hàng",
        debt_column="Nợ cần thu hiện tại",
    )


def generate_customer_codes(df: pd.DataFrame) -> pd.DataFrame:
    """Generate unified customer codes (KH000001, KH000002, ...)."""
    return generate_entity_codes(
        df=df,
        name_column="Tên khách hàng",
        code_column="Mã khách hàng",
        code_prefix="KH",
        date_column="first_date",
        amount_column="total_amount",
    )


def map_to_kiotviet_template(df: pd.DataFrame) -> pd.DataFrame:
    """Map to KiotViet 20-column customer template."""
    if df.empty:
        return df

    template = CustomerTemplate()
    columns = template.get_column_names()

    result = pd.DataFrame(index=range(len(df)), columns=columns)
    result = result.fillna("")

    result["Loại khách"] = "Cá nhân"
    result.loc[: len(df) - 1, "Mã khách hàng"] = df["Mã khách hàng"].values
    result.loc[: len(df) - 1, "Tên khách hàng"] = df["Tên khách hàng"].values

    if "Điện thoại" in df.columns:
        phones = df["Điện thoại"].apply(
            lambda x: split_phone_numbers(x)[0] if split_phone_numbers(x) else ""
        )
        result.loc[: len(df) - 1, "Điện thoại"] = phones.astype(str).values

    if "Địa chỉ" in df.columns:
        result.loc[: len(df) - 1, "Địa chỉ"] = df["Địa chỉ"].fillna("").values

    if "Ghi chú" in df.columns:
        ghi_chu = df["Ghi chú"].fillna("")
        ma_kh_cu = df.get("Mã KH cũ", pd.Series([""] * len(df)))
        if isinstance(ma_kh_cu, str):
            ma_kh_cu = pd.Series([ma_kh_cu] * len(df))
        ma_kh_cu = ma_kh_cu.fillna("")
        combined = []
        for gh, mk in zip(ghi_chu, ma_kh_cu):
            parts = []
            if gh and str(gh).strip():
                parts.append(str(gh))
            if mk and str(mk).strip():
                parts.append(f"Mã cũ: {mk}")
            combined.append("\n".join(parts))
        result.loc[: len(df) - 1, "Ghi chú"] = combined

    if "last_date" in df.columns:
        dates = pd.to_datetime(df["last_date"], errors="coerce").dt.strftime("%Y-%m-%d")
        dates = dates.replace("NaT", "")
        result.loc[: len(df) - 1, "Ngày giao dịch cuối"] = dates.values

    if "Nợ cần thu hiện tại" in df.columns:
        result.loc[: len(df) - 1, "Nợ cần thu hiện tại"] = (
            pd.to_numeric(df["Nợ cần thu hiện tại"], errors="coerce").fillna(0).values
        )

    if "total_amount" in df.columns:
        result.loc[: len(df) - 1, "Tổng bán (Không import)"] = (
            pd.to_numeric(df["total_amount"], errors="coerce").fillna(0).values
        )

    result.loc[: len(df) - 1, "Nhóm khách hàng"] = ""

    if "last_date" in df.columns and "Nợ cần thu hiện tại" in df.columns:
        from datetime import datetime, timedelta

        cutoff_date = datetime.now() - timedelta(days=90)
        last_date = pd.to_datetime(df["last_date"], errors="coerce")
        debt = pd.to_numeric(df["Nợ cần thu hiện tại"], errors="coerce").fillna(0)

        has_recent_transaction = (last_date >= cutoff_date) & last_date.notna()
        has_debt = debt > 0

        is_active = has_recent_transaction | has_debt
        result.loc[: len(df) - 1, "Trạng thái"] = is_active.apply(
            lambda x: 1 if x else 0
        ).values
    else:
        result.loc[: len(df) - 1, "Trạng thái"] = 1

    result = result.infer_objects(copy=False)

    # Fill NaN values per column based on data type
    template = CustomerTemplate()
    for col_spec in template.COLUMNS:
        col_name = col_spec.name
        if col_name in result.columns:
            if col_spec.data_type == "date":
                result[col_name] = result[col_name].fillna(pd.NaT)
            else:
                result[col_name] = result[col_name].fillna("")

    result["Điện thoại"] = result["Điện thoại"].astype(str)

    return result


def write_xlsx(df: pd.DataFrame, output_path: Path) -> None:
    """Write DataFrame to XLSX with formatting."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    template = CustomerTemplate()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Khách hàng"

    for col_idx, col_spec in enumerate(template.COLUMNS, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=col_spec.name)

    for idx, row in df.iterrows():
        for col_idx, col_spec in enumerate(template.COLUMNS, start=1):
            value = row[col_spec.name]
            if pd.isna(value):
                value = ""
            elif col_spec.data_type == "date":
                if pd.notna(value) and isinstance(value, pd.Timestamp):
                    value = value.to_pydatetime()
                elif pd.notna(value) and not isinstance(value, str):
                    value = str(value)
            elif not isinstance(value, str):
                value = str(value)
            cell = worksheet.cell(row=idx + 2, column=col_idx, value=value)
            if col_spec.data_type == "text":
                cell.number_format = "@"

    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx in range(1, len(template.COLUMNS) + 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, col_spec in enumerate(template.COLUMNS, start=1):
        letter = worksheet.cell(row=1, column=col_idx).column_letter
        worksheet.column_dimensions[letter].width = 20

        if col_spec.format_code and col_spec.data_type == "number":
            for row in range(2, len(df) + 2):
                cell = worksheet.cell(row=row, column=col_idx)
                cell.number_format = col_spec.format_code
                cell.alignment = Alignment(horizontal="right")
        elif col_spec.format_code and col_spec.data_type == "date":
            for row in range(2, len(df) + 2):
                cell = worksheet.cell(row=row, column=col_idx)
                cell.number_format = col_spec.format_code

    workbook.save(output_path)
    logger.info(f"Wrote XLSX: {output_path}")


def format_value(value, data_type: str) -> any:
    """Format value based on data type for Google Sheets."""
    if value is None or value == "" or (isinstance(value, float) and pd.isna(value)):
        return ""

    if data_type == "number":
        if isinstance(value, (int, float)):
            return value
        try:
            cleaned = str(value).replace(",", "").replace(".", "").replace(" ", "")
            if cleaned.startswith("(") and cleaned.endswith(")"):
                cleaned = "-" + cleaned[1:-1]
            return float(cleaned)
        except (ValueError, TypeError):
            return str(value)

    elif data_type == "date":
        if pd.isna(value) or value is None or value == "":
            return ""
        if isinstance(value, str):
            if "-" in value and len(value) >= 10:
                return value[:10]
            try:
                parsed = pd.to_datetime(value, errors="coerce")
                if pd.notna(parsed):
                    return parsed.strftime("%Y-%m-%d")
            except Exception:
                pass
        elif isinstance(value, pd.Timestamp):
            try:
                return value.strftime("%Y-%m-%d")
            except Exception:
                pass
        elif hasattr(value, "strftime"):
            try:
                return value.strftime("%Y-%m-%d")
            except Exception:
                pass
        return str(value) if value else ""

    else:
        return str(value)


def upload_to_google_sheet(df: pd.DataFrame, sheets_service) -> bool:
    """Upload DataFrame to Google Sheet as tab with proper formatting."""
    logger.info(
        f"Uploading to Google Sheet: {EXPORT_SPREADSHEET_ID}/{EXPORT_SHEET_NAME}"
    )

    template = CustomerTemplate()
    col_types = {col.name: col.data_type for col in template.COLUMNS}

    header = df.columns.tolist()
    rows = []

    for _, row in df.iterrows():
        formatted_row = []
        for col in header:
            data_type = col_types.get(col, "text")
            formatted_row.append(format_value(row[col], data_type))
        rows.append(formatted_row)

    template_df = pd.DataFrame(rows, columns=header)

    success = upload_dataframe_to_sheet(
        sheets_service,
        EXPORT_SPREADSHEET_ID,
        EXPORT_SHEET_NAME,
        template_df,
        raw_columns=[1],
    )

    if success:
        logger.info(f"Successfully uploaded {len(df)} rows to {EXPORT_SHEET_NAME}")
    else:
        logger.error("Failed to upload to Google Sheet")

    return success


def process(staging_dir: Optional[Path] = None) -> Optional[Path]:
    """Main processing function.

    Args:
        staging_dir: Optional path to staging directory (unused, for API consistency).
                    This function reads from Google Sheets directly.

    Returns:
        Path to generated Customers.xlsx or None if failed.
    """
    logger.info("=" * 70)
    logger.info("GENERATING CUSTOMERS XLSX (V2 - Single Script)")
    logger.info("=" * 70)

    DATA_EXPORT_DIR.mkdir(parents=True, exist_ok=True)

    try:
        creds = authenticate_google()
        from googleapiclient.discovery import build

        sheets_service = build("sheets", "v4", credentials=creds)

        customers_df = load_thong_tin_kh(sheets_service)
        debts_df = load_tong_cong_no(sheets_service)
        transactions_df = aggregate_transactions(load_sale_transactions())

        merged_df = merge_all_data(customers_df, debts_df, transactions_df)

        if merged_df.empty:
            logger.error("No customer data to export")
            return None

        merged_df = generate_customer_codes(merged_df)
        template_df = map_to_kiotviet_template(merged_df)

        output_path = DATA_EXPORT_DIR / "Customers.xlsx"
        write_xlsx(template_df, output_path)

        upload_to_google_sheet(template_df, sheets_service)

        logger.info(f"Generated {len(template_df)} customer records")
        logger.info(f"Output: {output_path}")
        logger.info("=" * 70)

        return output_path

    except Exception as e:
        logger.error(f"Failed to generate customers XLSX: {str(e)}", exc_info=True)
        raise


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
    )
    process()
